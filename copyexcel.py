#coding=utf-8
from openpyxl import load_workbook

#打开excel文件
wbhz = load_workbook(filename='附件1：1000号工单收集（汇总）.xlsx',data_only=False)
wbbq = load_workbook(filename='附件1：第2个页签汇总表（汇总）.xlsx',data_only=False)
#打开sheet
sheet = wbhz['6+1系统分析']
sheet1 = wbhz['1000号工单汇总']
#设置要打开的excel表格与sheet
excelname = ['附件1：1000号工单收集（总部）.xlsx','附件1：1000号工单收集（超高压）.xlsx','附件1：1000号工单收集（调峰调频）.xlsx','附件1：1000号工单收集（广东电网）.xlsx','附件1：1000号工单收集（广西电网）.xlsx','附件1：1000号工单收集（云南电网）.xlsx','附件1：1000号工单收集（贵州电网）.xlsx','附件1：1000号工单收集（海南电网）.xlsx','附件1：1000号工单收集（广州电网）.xlsx','附件1：1000号工单收集（深圳电网）.xlsx']
sheetname = ['总部','超高压','调峰调频','广东','广西','云南','贵州','海南','广州','深圳'] 
labelsheetname = ['（总部）','（超高压）','（双调）','（广东）','（广西）','（云南）','（贵州）','（海南）','（广州局）','（深圳局）']

#因第三方库openpyxl模式问题，会导致公式丢失或者数据丢失，功能暂时开放，调试好在开放
# i = ['3','4','5','6','7','8','9','10']
# for i1 in i:
#     s = sheet['J'+i1].value
#     sheet['K'+i1].value = s  
# print('已将6+1系统分析本期数据复制至上期数据')
# 
# j=['3','4','5','6','7','8','9','10','11','12','13','14']
# j1=['I','M','Q','U','Y','AC']
# j2=['J','N','R','V','Z','AD']
# for ja in range(len(j)):
#     for ja1 in range(len(j1)):
#         a = sheet1[j1[ja1]+j[ja]].value
#         sheet1[j2[ja1]+j[ja]].value = a
# print('已将1000号工单汇总本期数据复制至上期数据')            


#将各单位1000号工单汇总至汇总表
def thousand(excelname,sheetname):
    wb1 = load_workbook(filename=excelname,data_only=True)
    #设定需要循环的列数
    z = ['A','B','C','D','E','F','G','H','I','J','K','L']
    #从1到120循环，并把每次循环的值赋值给h
    for h in range(1,120):
        #循环z的长度，并把每次循环得到的值赋值给z1
        for z1 in range(len(z)):
            #设定打开对应的sheet(sheetname)
            sheet3 = wbhz[sheetname]
            sheet2 = wb1['1000号工单收集']
            #把seet2的值赋值给q
            q = sheet2[z[z1]+str(h)].value
            sheet3[z[z1]+str(h)].value = q
    print(sheetname+'已汇总')
              
#循环打开表格并赋值
for collect in range(len(excelname)):
    thousand(excelname[collect], sheetname[collect])

#将各单位第二页标签做汇总
def label(excelname,labelsheetname): 
    z = ['A','B','C','D','E','F','G','H','I','J','K','L'] 
    wb2 = load_workbook(filename=excelname,data_only=True)
    for h in range(1,120):
        for z1 in range(len(z)):
            sheet3 = wb2['1000号数据统计分析']
            sheet2 = wbbq[labelsheetname]
            q = sheet3[z[z1]+str(h)].value
            sheet2[z[z1]+str(h)].value = q
    print(labelsheetname+'已汇总')

for collect in range(len(excelname)):
    label(excelname[collect], labelsheetname[collect])

#保存 
wbhz.save('附件1：1000号工单收集（汇总）.xlsx')
wbbq.save('附件1：第2个页签汇总表（汇总）.xlsx')
wbhz.close()
wbbq.close()
print('已汇总完毕') 

#把信息汇总至重要表内
significance = load_workbook(filename='新模板表格数据（重要）.xlsx',data_only=False)
wbbq = load_workbook(filename='附件1：第2个页签汇总表（汇总）.xlsx',data_only=True)
wbhz = load_workbook(filename='附件1：1000号工单收集（汇总）.xlsx',data_only=True)
#取重要数据同步word数据源文件的表一
significanceEnglish = ['G','H','I','J','K','L','M']
significanceFigure = ['3','4','5','7','8','10']
significanceFigureone = ['11','12','13','15']
#将同步word数据表复制到旁边的表
significanCecopy = ['C','D','E','F','G','H','I','J','K','L','M']
significanFigureCecopy = ['3','4','5','6','7','8','9','10','11','12','13','14','15','16']
significanCecopyTarget = ['Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ']
significanFigureCecopyTarget = ['3','4','5','6','7','8','9','10','11','12','13','14','15','16']
#取1000号表与第二页标签表的内容，赋值给重要表
wbbqEnglish = ['D','E','F','G','H','I','J']
wbbqFigure = ['31','32','34','37','38','35']
thousandEnglish = ['E','F','G','H','I']
thousandFigure = ['3','4','5','6','7','8','9']
#将第二呀缺陷统计情况表赋值给重要表
wbbqDefect = ['B','C','D','E','F','G','H']
wbbqDefectFigure = ['57','58','59','60','61','62','63']
significanceFigure = ['4','5','6','7','8','9','10']
#重要数据缺陷统计情况表数据赋值
wbbqDefectFigureCopy = ['17','18','19','20','21','22','23']
wbbqDefectOne = ['B','C','D','E','H']
wbbqDefectTwo = ['K','L','M','N','O']




